import glob
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
#SN_FOLDER   = '/snpy_txt_clean/'
OUTPUT_FILE = 'snoopy_results.xlsx'

PARAMETERS = ['zcmb', 'DM', 'DM_err', 'dm15', 'dm15_err', 'Tmax', 'Tmax_err']

# ── Setup ─────────────────────────────────────────────────────────────────────
#the_list    = glob.glob(SN_FOLDER + 'SN*.txt')
SN_list = ['2025acsn', '2019np','2020ue', '2025aebt', '2025afwg', '2025ahkt', '2025ahqr', '2025ahsa', '2026acd', '2026atb', '2026gm', '2026kc', '2026dix', '2026kc_subbed', '2025acsn_subbed' ]
the_list = [SN+'.txt' for SN in SN_list]
print(the_list)
all_results = {}
fout        = open('failures.log', 'w')

# ── Main loop (CSP Example 2 pattern) ────────────────────────────────────────
for file in the_list:
    try:
        s = get_sn(file)                          # no snpy. prefix in IPython shell
        s.choose_model('EBV_model2', stype='dm15')
        s.fit()  
        s.get_zcmb()
        zcmb = s.zcmb                                 # fit all bands, no priors

        all_results[s.name] = {
            'zcmb'    : zcmb,
            'DM'      : s.parameters.get('DM',   None),
            'DM_err'  : s.errors.get('DM',       None),
            'dm15'    : s.parameters.get('dm15',  None),
            'dm15_err': s.errors.get('dm15',      None),
            'Tmax'    : s.parameters.get('Tmax',  None),
            'Tmax_err': s.errors.get('Tmax',      None),
        }
        s.save(s.name + '.snpy')                  # save fit for later inspection
        print(f'  ✓  {s.name} fitted successfully')

    except:
        fout.write('Failed on object %s\n' % s.name)
        all_results[s.name] = {p: None for p in PARAMETERS}

fout.close()

# ── Build Excel output ────────────────────────────────────────────────────────
param_labels = {
    'zcmb'    : 'z_CMB',
    'DM'      : 'DM (mag)',
    'DM_err'  : 'DM error (mag)',
    'dm15'    : 'Δm₁₅ (mag)',
    'dm15_err': 'Δm₁₅ error (mag)',
    'Tmax'    : 'T_max (MJD)',
    'Tmax_err': 'T_max error (days)',
}

header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill('solid', start_color='2E4057')
param_font   = Font(name='Arial', bold=True, size=10)
param_fill   = PatternFill('solid', start_color='D9E1F2')
data_font    = Font(name='Arial', size=10)
failed_fill  = PatternFill('solid', start_color='FFD7D7')
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin_side    = Side(style='thin', color='BFBFBF')
thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

failed_SNe = [name for name, vals in all_results.items() if all(v is None for v in vals.values())]
sn_names   = list(all_results.keys())

wb = Workbook()
ws = wb.active
ws.title = 'SNooPy Results'

# Header row
ws.cell(row=1, column=1, value='Parameter').font  = header_font
ws.cell(row=1, column=1).fill                      = header_fill
ws.cell(row=1, column=1).alignment                 = center_align
ws.cell(row=1, column=1).border                    = thin_border

for col_idx, sn_name in enumerate(sn_names, start=2):
    cell = ws.cell(row=1, column=col_idx, value=sn_name)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border

# Parameter rows
for row_idx, param in enumerate(PARAMETERS, start=2):
    cell = ws.cell(row=row_idx, column=1, value=param_labels[param])
    cell.font = param_font; cell.fill = param_fill
    cell.alignment = left_align; cell.border = thin_border

    for col_idx, sn_name in enumerate(sn_names, start=2):
        value = all_results[sn_name].get(param, None)
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font; cell.alignment = center_align
        cell.border = thin_border; cell.number_format = '0.0000'
        if sn_name in failed_SNe:
            cell.fill = failed_fill

ws.column_dimensions['A'].width = 22
for col_idx in range(2, len(sn_names) + 2):
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
ws.freeze_panes = 'B2'

wb.save(OUTPUT_FILE)
print(f'\nResults written to {OUTPUT_FILE}')
print(f'Check failures.log for any problematic SNe ({len(failed_SNe)} failed)')