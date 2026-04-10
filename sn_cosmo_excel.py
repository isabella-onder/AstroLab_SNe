import sncosmo
from astropy.table import Table
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
alpha = 0.148 
beta  = 3.112 #(Pantheon+, Brout et al. 2022)
M0    =  -19.253 #for H₀ = 73.04 km s⁻¹ Mpc⁻¹ (SH0ES, Riess et al. 2022)

SN_list = ['2019np', '2020ue', '2025ahsa', '2026acd', '2026atb','2026gm', '2026kc_subbed', '2025acsn_subbed']
z_cmb   = [0.0055, 0.0041, 0.0158, 0.0087, 0.0186, 0.0182, 0.0240, 0.0148]

PARAMETERS = ['z_cmb', 't0', 'x0', 'x1', 'colour', 'mB', 'mu',
              'chi2', 'ndof', 'chi2_red']

# ── Load full data table ──────────────────────────────────────────────────────
# assumes your table has a column called 'name' or 'sn' identifying each SN
# adjust 'name' below to match your actual column name
SN_COL = 'name'

all_results = {}
failures    = []

model = sncosmo.Model(source='salt2')

for SN, z in zip(SN_list, z_cmb):
    print(f'\n── Fitting {SN} (z_CMB = {z}) ──')
    try:
        # Read this SN's own input file
        data = Table.read(f"sncosmo/{SN}_input.dat", format="ascii")

        model.set(z=z)

        result, fitted_model = sncosmo.fit_lc(
            data, model,
            ['t0', 'x0', 'x1', 'c'],
            bounds={'x1': (-5, 5), 'c': (-1, 2)}
        )

        t0_fit = result.parameters[1]
        x0_fit = result.parameters[2]
        x1_fit = result.parameters[3]
        c_fit  = result.parameters[4]

        mB     = -2.5 * np.log10(x0_fit) + 10.635
        mu     = mB + alpha * x1_fit - beta * c_fit - M0
        chi2_red = result.chisq / result.ndof if result.ndof > 0 else None

        all_results[SN] = {
            'z_cmb'   : z,
            't0'      : t0_fit,
            'x0'      : x0_fit,
            'x1'      : x1_fit,
            'colour'  : c_fit,
            'mB'      : mB,
            'mu'      : mu,
            'chi2'    : result.chisq,
            'ndof'    : result.ndof,
            'chi2_red': chi2_red,
        }
        print(f'  ✓  mu = {mu:.4f},  chi2/ndof = {chi2_red:.3f}')

    except Exception as e:
        print(f'  ✗  FAILED: {e}')
        failures.append(SN)
        all_results[SN] = {p: None for p in PARAMETERS}

# ── Excel output ──────────────────────────────────────────────────────────────
param_labels = {
    'z_cmb'   : 'z_CMB',
    't0'      : 'T₀ (MJD)',
    'x0'      : 'x₀',
    'x1'      : 'x₁',
    'colour'  : 'c (colour)',
    'mB'      : 'mB (mag)',
    'mu'      : 'μ (mag)',
    'chi2'    : 'χ²',
    'ndof'    : 'N_dof',
    'chi2_red': 'χ²/N_dof',
}

header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill('solid', start_color='2E4057')
param_font   = Font(name='Arial', bold=True, size=10)
param_fill   = PatternFill('solid', start_color='D9E1F2')
data_font    = Font(name='Arial', size=10)
failed_fill  = PatternFill('solid', start_color='FFD7D7')
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin_side    = Side(style='thin', color='BFBFBF')
thin_border  = Border(left=thin_side, right=thin_side,
                      top=thin_side,  bottom=thin_side)

wb = Workbook()
ws = wb.active
ws.title = 'SALT2 Results'

# Header row — one column per SN
ws.cell(row=1, column=1, value='Parameter').font      = header_font
ws.cell(row=1, column=1).fill                          = header_fill
ws.cell(row=1, column=1).alignment                     = center_align
ws.cell(row=1, column=1).border                        = thin_border

for col_idx, sn in enumerate(SN_list, start=2):
    cell = ws.cell(row=1, column=col_idx, value=sn)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border

# One row per parameter
for row_idx, param in enumerate(PARAMETERS, start=2):
    cell = ws.cell(row=row_idx, column=1, value=param_labels[param])
    cell.font = param_font; cell.fill = param_fill
    cell.alignment = left_align; cell.border = thin_border

    for col_idx, sn in enumerate(SN_list, start=2):
        value = all_results[sn].get(param)
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font; cell.alignment = center_align
        cell.border = thin_border; cell.number_format = '0.00000'
        if sn in failures:
            cell.fill = failed_fill

ws.column_dimensions['A'].width = 18
for col_idx in range(2, len(SN_list) + 2):
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
ws.freeze_panes = 'B2'

wb.save('snpy_txt_clean/salt2_results.xlsx')
print(f'\nResults written to salt2_results.xlsx')
print(f'Failed SNe ({len(failures)}): {failures}')