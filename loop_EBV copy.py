import glob
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
OUTPUT_FILE = 'snoopy_results_EBV_fixed_max.xlsx'

PARAMETERS = ['zcmb', 'DM', 'DM_err', 'dm15', 'dm15_err', 'Tmax', 'Tmax_err', 'i_max', 'i_max_error', 'r_max', 'r_max_error','v_max', 'v_max_error', 'b_max', 'b_max_error']

# ── Optional EBVhost priors ───────────────────────────────────────────────────
# Must be in the same order as SN_list. Use None for a free fit on individual
# entries, or set the entire list to None to fit all SNe freely.
# e.g. EBV_host_list = [0.1, None, 0.3, None, 0.05, ...]

SN_list = ['2025acsn', '2019np', '2020ue', '2025aebt',  '2025ahkt',
           '2025ahqr', '2025ahsa', '2026acd', '2026atb', '2026gm', '2026kc',
           '2026dix', '2026kc_subbed', '2025acsn_subbed']

EBV_host_list = [0.1586433260393873, 0.020058351568198397, 0.028446389496717725, 0.07221006564551423, 0.38439095550692925,  0.016776075857038657, 0.04412837345003647, 0.03646973012399708, 0.05397520058351568, 0.3931436907366886, 0.03391684901531729, 0.01276440554339898, 0.03391684901531729, 0.1586433260393873]   # or e.g. [0.1, 0.2, None, 0.05, ...]

# ── Setup ─────────────────────────────────────────────────────────────────────
the_list = [SN + '.txt' for SN in SN_list]
print(the_list)

if EBV_host_list is not None and len(EBV_host_list) != len(SN_list):
    raise ValueError(
        f'EBV_host_list has {len(EBV_host_list)} entries but SN_list has '
        f'{len(SN_list)}. These must match exactly.'
    )

all_results = {}
fout        = open('failures.log', 'w')

# ── Main loop (CSP Example 2 pattern) ────────────────────────────────────────
for idx, file in enumerate(the_list):
    try:
        s = get_sn(file)
        s.choose_model('EBV_model2', stype='dm15')

        if EBV_host_list is not None and EBV_host_list[idx] is not None:
            ebv = EBV_host_list[idx]
            print(f'  Fitting {s.name} with EBVhost = {ebv} (fixed prior)')
            s.fit(EBVhost=ebv)
        else:
            print(f'  Fitting {s.name} with EBVhost free')
            s.fit()

        try:
            s.get_zcmb()
            zcmb = s.zcmb
            maxes = s.get_max(['Is', 'Rs', 'Vs', 'Bs'])
            print('this is maxes')
            i_max = maxes[1][0]
            i_max_error = maxes[2][0]
            r_max = maxes[1][1]
            r_max_error = maxes[2][1]
            v_max = maxes[1][2]
            v_max_error = maxes[2][2]
            b_max = maxes[1][3]
            b_max_error = maxes[2][3]

        except:
            zcmb = getattr(s, 'z', None)
            print(f'  ⚠  {s.name}: zcmb correction unavailable, using heliocentric z')

        all_results[s.name] = {
            'zcmb'    : zcmb,
            'DM'      : s.parameters.get('DM',    None),
            'DM_err'  : s.errors.get('DM',        None),
            'dm15'    : s.parameters.get('dm15',   None),
            'dm15_err': s.errors.get('dm15',       None),
            'Tmax'    : s.parameters.get('Tmax',   None),
            'Tmax_err': s.errors.get('Tmax',       None),
            'i_max' : i_max,
            'i_max_error' : i_max_error,
            'r_max' : r_max,
            'r_max_error' : r_max_error,
            'v_max' : v_max,
            'v_max_error' : v_max_error,
            'b_max' : b_max,
            'b_max_error' : b_max_error
        }
        s.save(s.name + '.snpy')
        print(f'  ✓  {s.name} fitted successfully')

    except:
        fout.write('Failed on object %s\n' % s.name)
        all_results[s.name] = {p: None for p in PARAMETERS}
        print('it has failed')

fout.close()

# ── Build Excel output ────────────────────────────────────────────────────────
param_labels = {
    'zcmb'    : 'z_CMB',
    'DM'      : 'DM (mag)',
    'DM_err'  : 'DM error (mag)',
    'dm15'    : 'Δm₁₅ (mag)',
    'dm15_err': 'Δm₁₅ error (mag)',
    'Tmax'    : 'T_max (MJD)',
    'Tmax_err': 'T_max error (days)',
    'i_max': 'i_max (mag)',
    'i_max_error': 'i_max_error (mag)',
    'r_max': 'r_max (mag)',
    'r_max_error': 'r_max_error (mag)',
    'v_max': 'v_max (mag)',
    'v_max_error': 'v_max_error (mag)',
    'b_max': 'v_max (mag)',
    'b_max_error': 'v_max_error (mag)',
}

header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill('solid', start_color='2E4057')
param_font   = Font(name='Arial', bold=True, size=10)
param_fill   = PatternFill('solid', start_color='D9E1F2')
data_font    = Font(name='Arial', size=10)
failed_fill  = PatternFill('solid', start_color='FFD7D7')
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin_side    = Side(style='thin', color='BFBFBF')
thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

failed_SNe = [name for name, vals in all_results.items() if vals.get('DM') is None]
sn_names   = list(all_results.keys())

wb = Workbook()
ws = wb.active
ws.title = 'SNooPy Results'

# Header row
ws.cell(row=1, column=1, value='Parameter').font  = header_font
ws.cell(row=1, column=1).fill                      = header_fill
ws.cell(row=1, column=1).alignment                 = center_align
ws.cell(row=1, column=1).border                    = thin_border

for col_idx, sn_name in enumerate(sn_names, start=2):
    cell = ws.cell(row=1, column=col_idx, value=sn_name)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = center_align; cell.border = thin_border

# Parameter rows
for row_idx, param in enumerate(PARAMETERS, start=2):
    cell = ws.cell(row=row_idx, column=1, value=param_labels[param])
    cell.font = param_font; cell.fill = param_fill
    cell.alignment = left_align; cell.border = thin_border

    for col_idx, sn_name in enumerate(sn_names, start=2):
        value = all_results[sn_name].get(param, None)
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font; cell.alignment = center_align
        cell.border = thin_border; cell.number_format = '0.0000'
        if sn_name in failed_SNe:
            cell.fill = failed_fill

ws.column_dimensions['A'].width = 22
for col_idx in range(2, len(sn_names) + 2):
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
ws.freeze_panes = 'B2'

wb.save(OUTPUT_FILE)
print(f'\nResults written to {OUTPUT_FILE}')
print(f'Check failures.log for any problematic SNe ({len(failed_SNe)} failed)')