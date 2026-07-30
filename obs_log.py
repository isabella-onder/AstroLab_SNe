"""
observation_table.py
--------------------
Reads a folder of CSV files (one per supernova, filename = SN name) and
produces a formatted Excel workbook with one sheet per telescope.

Layout per sheet:
    rows    → unique observation dates (calendar date, sorted)
    col 1   → Date
    col 2+  → one column per supernova

Cell content: coloured ✗ symbol(s) for each band observed on that date.
    B → blue   (#4472C4)
    V → green  (#70AD47)
    R → red    (#FF0000)
    I → yellow (#FFD700)

Multiple bands on the same night appear as stacked coloured ✗ symbols
in separate lines within the same cell.

Usage:
    python observation_table.py --folder /path/to/csvs --output obs_table.xlsx
"""
Photometry_final_uncleaned_backup
import os
import glob
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock

# ── Band → hex colour ──────────────────────────────────────────────────────
BAND_COLOURS = {
    "B": "4472C4",   # blue
    "V": "70AD47",   # green
    "R": "FF0000",   # red
    "I": "FFD700",   # yellow / gold
}
BAND_ORDER = ["B", "V", "R", "I"]
MARKER = "✗"

# ── Styles ─────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="2F3640")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATE_FONT     = Font(name="Arial", size=9, color="2C2C2C")
DATE_FILL_ODD = PatternFill("solid", fgColor="F5F5F5")
DATE_FILL_EVN = PatternFill("solid", fgColor="FFFFFF")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center")
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_csvs(folder):
    frames = []
    for path in sorted(glob.glob(os.path.join(folder, "*.csv"))):
        sn_name = os.path.splitext(os.path.basename(path))[0]
        df = pd.read_csv(path)
        df.columns = df.columns.str.strip()
        df["date_parsed"] = pd.to_datetime(
            df["date"], errors="coerce"
        ).dt.date
        df["SN"] = sn_name
        frames.append(df[["SN", "Telescope", "date_parsed", "Filter"]])
    if not frames:
        raise FileNotFoundError(f"No CSV files found in {folder!r}")
    return pd.concat(frames, ignore_index=True)


def build_tables(df):
    tables = {}
    for telescope, tdf in df.groupby("Telescope"):
        mapping = {}
        for _, row in tdf.iterrows():
            key = (row["date_parsed"], row["SN"])
            mapping.setdefault(key, [])
            band = str(row["Filter"]).strip().upper()
            if band in BAND_COLOURS and band not in mapping[key]:
                mapping[key].append(band)
        for key in mapping:
            mapping[key] = [b for b in BAND_ORDER if b in mapping[key]]
        tables[telescope] = {
            "mapping": mapping,
            "dates": sorted({k[0] for k in mapping}),
            "sne":   sorted({k[1] for k in mapping}),
        }
    return tables


def make_rich_string(bands):
    parts = []
    for i, band in enumerate(bands):
        colour = BAND_COLOURS[band]
        ifont  = InlineFont(color=colour, b=True, sz=12, rFont="Arial")
        text   = MARKER if i == 0 else f"\n{MARKER}"
        parts.append(TextBlock(ifont, text))
    return CellRichText(*parts)


def write_sheet(wb, telescope, info):
    ws = wb.create_sheet(title=telescope[:31])
    dates   = info["dates"]
    sne     = info["sne"]
    mapping = info["mapping"]

    # Header
    for j, label in enumerate(["Date"] + sne, start=1):
        c = ws.cell(1, j, label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = BORDER

    # Data
    for i, date in enumerate(dates, start=2):
        row_fill = DATE_FILL_ODD if i % 2 else DATE_FILL_EVN
        dc = ws.cell(i, 1, str(date))
        dc.font      = DATE_FONT
        dc.fill      = row_fill
        dc.alignment = LEFT
        dc.border    = BORDER

        for j, sn in enumerate(sne, start=2):
            c = ws.cell(i, j)
            c.fill      = row_fill
            c.alignment = CENTER
            c.border    = BORDER
            bands = mapping.get((date, sn), [])
            if bands:
                c.value = make_rich_string(bands)

    # Column widths
    ws.column_dimensions["A"].width = 13
    for j, sn in enumerate(sne, start=2):
        ws.column_dimensions[get_column_letter(j)].width = max(10, len(sn) + 2)

    # Row heights
    ws.row_dimensions[1].height = 20
    for i, date in enumerate(dates, start=2):
        max_bands = max(
            (len(mapping.get((date, sn), [])) for sn in sne), default=1
        )
        ws.row_dimensions[i].height = max(15, max_bands * 15)

    # Legend
    legend_row = len(dates) + 3
    lc = ws.cell(legend_row, 1, "Legend:")
    lc.font = Font(bold=True, name="Arial", size=9)
    for k, band in enumerate(BAND_ORDER, start=2):
        lc2 = ws.cell(legend_row, k, f"{MARKER}  {band}")
        lc2.font      = Font(color=BAND_COLOURS[band], bold=True, name="Arial", size=10)
        lc2.alignment = CENTER


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--folder",  default=".",              help="Folder containing CSV files")
    parser.add_argument("--output",  default="obs_table.xlsx", help="Output Excel filename")
    args = parser.parse_args()

    print(f"Reading CSVs from: {args.folder}")
    df = load_csvs(args.folder)
    print(f"  {len(df)} rows | {df['SN'].nunique()} SN(e) | {df['Telescope'].nunique()} telescope(s)")

    tables = build_tables(df)
    wb = Workbook()
    wb.remove(wb.active)

    for telescope in sorted(tables):
        print(f"  Writing sheet: {telescope}")
        write_sheet(wb, telescope, tables[telescope])

    wb.save(args.output)
    print(f"Saved → {args.output}")


if __name__ == "__main__":
    main()
